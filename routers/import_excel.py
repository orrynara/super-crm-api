from fastapi import APIRouter, UploadFile, File, HTTPException, Request
from typing import Optional
import pandas as pd
import io
from services.supabase_client import supabase
from models.schemas import ApiResponse

router = APIRouter(prefix="/api/v1/import", tags=["import"])


# ── 고객등급 → category 매핑 ─────────────────────────
def map_category(grade: str) -> str:
    if not grade or pd.isna(grade):
        return "new"
    g = str(grade).strip()
    if "01" in g or "우수" in g or "VIP" in g.upper() or "인지" in g:
        return "vip"
    if "02" in g or "기왕" in g or "기존" in g:
        return "existing"
    if "03" in g or "관리" in g:
        return "existing"
    if "04" in g or "DB" in g.upper() or "신규" in g:
        return "new"
    if "05" in g or "휴면" in g:
        return "dormant"
    if "06" in g or "하루" in g or "가족" in g or "협력" in g:
        return "existing"
    return "new"


# ── 생년월일 조합 (4자리 년도 우선) ──────────────────
def build_birth_date(year_val, month_val, day_val) -> Optional[str]:
    try:
        y = str(year_val).strip().replace(".0", "")
        m = str(month_val).strip().replace(".0", "")
        d = str(day_val).strip().replace(".0", "")
        if not y or y in ["nan", "None", ""]:
            return None
        yy = int(float(y))
        mm = int(float(m)) if m and m not in ["nan", "None", ""] else 0
        dd = int(float(d)) if d and d not in ["nan", "None", ""] else 0
        if mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return None
        # 2자리 년도 처리 (주민번호 년)
        if yy < 100:
            yy = 2000 + yy if yy <= 24 else 1900 + yy
        return f"{yy:04d}-{mm:02d}-{dd:02d}"
    except Exception:
        return None


# ── 주민번호 앞자리 → 생년월일 변환 ─────────────────
def parse_birth(resident_no) -> Optional[str]:
    try:
        s = str(resident_no).replace("-", "").replace(".", "").strip()
        s = s[:6]
        if len(s) < 6:
            return None
        yy = int(s[0:2])
        mm = int(s[2:4])
        dd = int(s[4:6])
        if mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return None
        year = 2000 + yy if yy <= 24 else 1900 + yy
        return f"{year:04d}-{mm:02d}-{dd:02d}"
    except Exception:
        return None


# ── 전화번호 정규화 ───────────────────────────────────
def normalize_phone(phone) -> Optional[str]:
    if not phone or pd.isna(phone):
        return None
    p = str(phone).replace(" ", "").replace("-", "")
    digits = "".join(c for c in p if c.isdigit())
    if len(digits) == 11 and digits.startswith("010"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return str(phone).strip() if str(phone).strip() else None


def _v(val) -> str:
    """pandas 셀 값 → 안전한 문자열"""
    s = str(val).strip()
    return "" if s in ["nan", "None", "NaT", ""] else s


# ── 컬럼 매핑 탐지 ───────────────────────────────────
def detect_col_map(columns: list) -> dict:
    m = {}
    for col in columns:
        c = col.strip()
        if c in ["이름", "성명", "고객이름", "name", "고객명"]:
            m["name"] = col
        elif c in ["휴대폰", "전화", "연락처", "핸드폰", "휴대전화", "phone", "전화번호", "연락처(휴대)"]:
            m["phone"] = col
        elif c in ["고객등급", "등급", "고객구분", "category", "구분"]:
            m["category"] = col
        elif c in ["주민번호앞", "주민앞", "생년월일", "birth", "주민번호", "생년월일(주민앞)"]:
            m["birth"] = col
        # 내고객다보여 분리형 생년월일 (생년/생월/생일 우선)
        elif c == "생년":
            m["birth_year"] = col
        elif c == "생월":
            m["birth_month"] = col
        elif c == "생일":
            m["birth_day"] = col
        # 주민번호 분리형 (백업)
        elif c == "주민번호 년":
            m["rr_year"] = col
        elif c == "주민번호 월":
            m["rr_month"] = col
        elif c == "주민번호 일":
            m["rr_day"] = col
        elif c in ["직업", "job"]:
            m["job"] = col
        elif c in ["성별", "gender"]:
            m["gender"] = col
        elif c in ["소개자", "introducer", "소개"]:
            m["introducer"] = col
        elif c in ["그룹분류", "그룹", "group"]:
            m["group"] = col
        # 집주소 분리형
        elif c in ["집주소 시/도", "시/도"]:
            m["addr_sido"] = col
        elif c in ["집주소 구/군/시", "구/군/시"]:
            m["addr_gu"] = col
        elif c in ["집주소 동/읍/로", "동/읍/로"]:
            m["addr_dong"] = col
        elif c in ["집주소 길/로", "길/로"]:
            m["addr_ro"] = col
        elif c in ["집주소 나머지", "나머지"]:
            m["addr_rest"] = col
        # 단일 주소
        elif c in ["주소", "집주소", "address", "거주지"]:
            m["address"] = col
        elif c in ["특이사항", "특이사항1", "메모", "memo", "비고"]:
            m["memo"] = col
    return m


# ── 행 데이터 → record dict ──────────────────────────
def parse_row(row, col_map: dict) -> Optional[dict]:
    name = _v(row.get(col_map["name"], ""))
    if not name:
        return None

    # 생년월일: 생년/생월/생일 우선 → 주민번호 년/월/일 → 통합 주민번호
    birth_date = None
    if "birth_year" in col_map and "birth_month" in col_map and "birth_day" in col_map:
        birth_date = build_birth_date(
            row.get(col_map["birth_year"], ""),
            row.get(col_map["birth_month"], ""),
            row.get(col_map["birth_day"], ""),
        )
    if not birth_date and "rr_year" in col_map and "rr_month" in col_map and "rr_day" in col_map:
        birth_date = build_birth_date(
            row.get(col_map["rr_year"], ""),
            row.get(col_map["rr_month"], ""),
            row.get(col_map["rr_day"], ""),
        )
    if not birth_date and "birth" in col_map:
        birth_date = parse_birth(row.get(col_map["birth"]))

    # 주소: 분리형 우선 → 단일 컬럼
    addr_parts = []
    for key in ["addr_sido", "addr_gu", "addr_dong", "addr_ro", "addr_rest"]:
        if key in col_map:
            v = _v(row.get(col_map[key], ""))
            if v:
                addr_parts.append(v)
    if addr_parts:
        address = " ".join(addr_parts)
    elif "address" in col_map:
        address = _v(row.get(col_map["address"], "")) or None
    else:
        address = None

    # 메모: 직업 / 성별 / 소개자 / 그룹
    memo_parts = []
    for key, label in [("job", "직업"), ("gender", "성별"), ("introducer", "소개자"), ("group", "그룹")]:
        if key in col_map:
            v = _v(row.get(col_map[key], ""))
            if v:
                memo_parts.append(f"{label}: {v}")

    return {
        "name":       name,
        "phone":      normalize_phone(row.get(col_map["phone"])) if "phone" in col_map else None,
        "category":   map_category(row.get(col_map["category"])) if "category" in col_map else "new",
        "status":     "active",
        "birth_date": birth_date,
        "address":    address,
        "memo":       " | ".join(memo_parts) or None,
    }


# ── 시트+헤더 행 자동 탐색 ──────────────────────────
def find_customer_sheet(contents: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    for sheet in sheet_names:
        for header_row in range(20):
            try:
                _df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet,
                                    header=header_row, engine="openpyxl")
                _df.columns = [str(c).strip() for c in _df.columns]
                col_map = detect_col_map(list(_df.columns))
                if "name" in col_map:
                    return _df, col_map, sheet
            except Exception:
                continue

    # 탐색 실패 시 디버그 정보 반환
    try:
        import openpyxl as ox
        wb2 = ox.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        snames = wb2.sheetnames
        wb2.close()
        samples = {}
        for s in snames[:3]:
            _df_raw = pd.read_excel(io.BytesIO(contents), sheet_name=s,
                                    header=None, engine="openpyxl")
            for i in range(min(5, len(_df_raw))):
                vals = [_v(v) for v in _df_raw.iloc[i].tolist() if _v(v)]
                if vals:
                    samples[f"{s}/row{i}"] = vals[:8]
        raise HTTPException(status_code=422,
            detail=f"'이름' 컬럼을 찾을 수 없습니다. 시트: {snames}, 샘플: {samples}")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=422, detail="'이름' 컬럼을 찾을 수 없습니다.")


@router.post("/excel/preview")
async def preview_excel(file: UploadFile = File(...)):
    """엑셀 파일 업로드 → 파싱 미리보기 (저장 안함)"""
    if not file.filename.endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(status_code=400, detail="Excel 파일(.xlsx, .xls, .xlsm)만 업로드 가능합니다.")
    try:
        contents = await file.read()
        df, col_map, found_sheet = find_customer_sheet(contents)

        records = []
        skipped = 0
        for _, row in df.iterrows():
            rec = parse_row(row, col_map)
            if rec is None:
                skipped += 1
            else:
                records.append(rec)

        return ApiResponse(
            success=True,
            data={
                "total":         len(records),
                "skipped":       skipped,
                "sheet":         found_sheet,
                "columns_found": list(col_map.keys()),
                "preview":       records[:10],
                "all":           records,
            },
            message=f"총 {len(records)}건 파싱 완료 (빈 행 {skipped}건 제외)"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 파싱 오류: {str(e)}")


@router.post("/excel/confirm")
async def confirm_import(request: Request, file: UploadFile = File(...)):
    """엑셀 파일 → Supabase 일괄 저장"""
    if not file.filename.endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(status_code=400, detail="Excel 파일만 업로드 가능합니다.")
    try:
        # JWT 토큰에서 agent_id 추출
        agent_id = None
        token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
        if token:
            try:
                user_resp = supabase.auth.get_user(token)
                if user_resp.user:
                    agent_id = str(user_resp.user.id)
            except Exception:
                pass

        contents = await file.read()
        df, col_map, _ = find_customer_sheet(contents)

        records = []
        skipped = 0
        for _, row in df.iterrows():
            rec = parse_row(row, col_map)
            if rec is None:
                skipped += 1
            else:
                # None 값 제거 후 저장, agent_id 자동 할당
                clean = {k: v for k, v in rec.items() if v is not None}
                if agent_id:
                    clean["agent_id"] = agent_id
                records.append(clean)

        if not records:
            raise HTTPException(status_code=422, detail="임포트할 유효한 데이터가 없습니다.")

        # Supabase bulk insert (100건씩 청크)
        inserted = 0
        errors = []
        for i in range(0, len(records), 100):
            chunk = records[i:i + 100]
            try:
                result = supabase.table("customers").insert(chunk).execute()
                inserted += len(result.data)
            except Exception as e:
                errors.append(f"청크 {i // 100 + 1} 오류: {str(e)[:100]}")

        return ApiResponse(
            success=True,
            data={"inserted": inserted, "skipped": skipped, "errors": errors},
            message=f"{inserted}건 임포트 완료" + (f" ({len(errors)}건 오류)" if errors else "")
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"임포트 오류: {str(e)}")
